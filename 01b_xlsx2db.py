#!/usr/bin/env python3
"""
Python replacement for 01b_xsl2db.R
Reads studies.xlsx and populates the project table in db.sqlite
"""
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
XLSX = BASE_DIR / "studies.xlsx"
DB = BASE_DIR / "db.sqlite"

# physicist name -> pitt email prefix
ADDR = {"Moon": "chm", "Schirda": "s", "Kim": "t"}


def pitt_email(name: str) -> str | None:
    name = name.strip()
    a = ADDR.get(name)
    return f"{a}@pitt.edu" if a else None


def to_emails(physicist: str) -> str:
    if not physicist:
        return ""
    # clean up physicist field like R script
    physicist = re.sub(r"(Victor|Yushmanov)[,/]?", "", physicist)
    physicist = re.sub(r"Moon\.2019\./", "", physicist)
    physicist = re.sub(r"\(.*?\)|set up", "", physicist)
    physicist = re.sub(r"/", ",", physicist)
    physicist = physicist.replace(" ", "")
    names = [n.strip() for n in physicist.split(",") if n.strip()]
    emails = [pitt_email(n) for n in names]
    return ",".join(e for e in emails if e)


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    col_map = {
        "PI Last Name": "PI",
        "MRRC Project Code": "Project",
        "Study Title": "Title",
        "Hrs": "Hrs",
        "Physicist": "Physicist",
        "PI Email": "email",
        "CC Email": "cc",
    }

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {col_map.get(h, h): v for h, v in zip(headers, row) if h in col_map}
        if not d.get("Project") or not d.get("PI"):
            continue
        d["Project"] = str(d["Project"]).upper().strip()
        physicist = str(d.get("Physicist", "") or "")
        d["phys_mail"] = to_emails(physicist)
        d["contact"] = physicist
        rows.append(d)

    sql = sqlite3.connect(str(DB))
    sql.execute("DROP TABLE IF EXISTS project")
    sql.execute(
        """
        CREATE TABLE project (
            PI TEXT,
            Project TEXT,
            Title TEXT,
            Hrs TEXT,
            Physicist TEXT,
            email TEXT,
            cc TEXT,
            contact TEXT,
            phys_mail TEXT
        )
    """
    )
    sql.executemany(
        """
        INSERT INTO project (PI, Project, Title, Hrs, Physicist, email, cc, contact, phys_mail)
        VALUES (:PI, :Project, :Title, :Hrs, :Physicist, :email, :cc, :contact, :phys_mail)
    """,
        rows,
    )
    sql.commit()
    print(f"Inserted {len(rows)} projects into project table")
    sql.close()


if __name__ == "__main__":
    main()
